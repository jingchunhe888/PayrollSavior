import pandas as pd
import numpy as np
from openpyxl import load_workbook



def time_total(data):
    # Define the pattern and replacement rules
    pattern = ['time in', 'time out', 'break', 'total']
    pattern_lower = [p.lower() for p in pattern]

    # Convert all data to lowercase for easier matching
    data_lower = [str(item).strip().lower() if pd.notna(item) else item for item in data]

    # Initialize the result list
    result_list = data.copy()

    i = 0
    while i < len(data_lower):
        if pd.notna(data[i]):
            current = data_lower[i]
            if current in pattern_lower:
                index = pattern_lower.index(current)
                if index == 0:  # 'time in'
                    # Replace the next 3 values
                    if i + 3 < len(result_list):
                        result_list[i + 1] = 'time out'
                        result_list[i + 2] = 'break'
                        result_list[i + 3] = 'total'
                elif index == 1:  # 'time out'
                    # Replace the previous value with 'time in', and the next 2 values
                    if i - 1 >= 0:
                        result_list[i - 1] = 'time in'
                    if i + 2 < len(result_list):
                        result_list[i + 1] = 'break'
                        result_list[i + 2] = 'total'
                elif index == 2:  # 'break'
                    # Replace the 2 values before with 'time in', 'time out', and 1 value after with 'total'
                    if i - 2 >= 0:
                        result_list[i - 2] = 'time in'
                    if i - 1 >= 0:
                        result_list[i - 1] = 'time out'
                    if i + 1 < len(result_list):
                        result_list[i + 1] = 'total'
        i += 1

    # Output the result
    return result_list

def nan_none(data_list):
    # Replace NaN values with None
    return [np.nan if pd.isna(item) else item for item in data_list]

#old merge rows
def merge_rows(df):
    # Create a copy of the DataFrame to work on
    df_copy = df.copy()
    df_columns = df.columns
    all_rows = []

    # Track rows to skip after merging
    skip_rows = set()

    i = 0
    num_merged_rows = 0
    while i < len(df_copy):
        if i in skip_rows:
            # If the row has already been merged, move to the next one
            i += 1
            continue

        current_row = df_copy.iloc[i].copy() # Create a copy of the current row
        if i + 1 < len(df_copy) - num_merged_rows:
            # print(f'Length of df {len(df_copy)}')
            next_row = df_copy.iloc[i + 1].copy() # Create a copy of the next row
        else: #it is the last row
            "pass"
            pass
            #next_row = pd.Series([np.nan] * len(df_copy.columns), index=df_copy.columns)

        merge = False
        skip_merge = False

        for col in range(len(df_copy.columns)):
            current_value = current_row.iloc[col]
            next_value = next_row.iloc[col]

            if (pd.notna(current_value)) & (pd.notna(next_value)):
                # Both rows have values in this column; skip merging
                skip_merge = True
                current_row_list = current_row.tolist()
                # print('I am he3434re1')
                # print(current_row_list)
                all_rows.append(current_row_list)
                break
            elif (pd.notna(current_value)) & (pd.isna(next_value)):
                # Current row has a value and next row has NaN
                merge = True
                break
            elif (pd.isna(current_value)) & (pd.notna(next_value)):
                # Current row has NaN and next row has a value
                merge = True
                break
            elif (pd.isna(current_value)) & (pd.isna(next_value)):
                # Both are NaN; we don't need to merge as no useful data is available
                merge = True
                break

        if merge and not skip_merge:
            # print('we are merging')
            num_merged_rows = num_merged_rows + 1
            # Merge the rows
            for col in range(len(df_copy.columns)):
                if (pd.isna(current_row.iloc[col])) and not (pd.isna(next_row.iloc[col])):
                    current_row.iloc[col] = next_row.iloc[col]
                # If both are NaN, keep NaN

            # Append the merged row to the result DataFrame
            current_row_list = current_row.tolist()
            # print('I am here1')
            # print(current_row_list)
            all_rows.append(current_row_list)
            # Mark the next row as skipped
            skip_rows.add(i + 1)
        else:
            pass
            # If not merging, just add the current row to the result DataFrame

            # current_row_list = current_row.tolist()
            # print('I am here22')
            # print(current_row_list)
            # all_rows.append(current_row_list)
        # print(f'all row')
        # print(all_rows)

        i += 1
    # print(f'final all row')
    # print(all_rows)
    df = pd.DataFrame(all_rows, columns=df_columns)
    # print(df.to_string())
    return df

def read_excel_exclude_hidden(file_path, sheet_name):
    # Load workbook without read_only mode
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]

    # Identify hidden columns
    hidden_columns = []
    for col in ws.column_dimensions:
        if ws.column_dimensions[col].hidden:
            hidden_columns.append(col)

    # Read excel with pandas
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Drop hidden columns from dataframe
    for col in hidden_columns:
        col_index = ws[col+'1'].column - 1  # Convert Excel column to zero-indexed column number
        if col_index < len(df.columns):
            df.drop(df.columns[col_index], axis=1, inplace=True)

    return df
