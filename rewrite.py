import csv
import os
import shutil
from openpyxl import load_workbook
import pandas as pd
def modify_cell(data, row_index, col_index, new_value):
    if row_index < len(data) and col_index < len(data[row_index]):
        data[row_index][col_index] = new_value
    else:
        # print(f"data length {len(data)}")
        # print(f"data length {len(data[5])}")
        print(f"Invalid row {row_index}or column index {col_index}.")

def read_csv(file_path):
    with open(file_path, mode='r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        data = list(reader)
        # for stuff in data:
        #     # print(stuff)
    return data

def write_csv(file_path, data):
    with open(file_path, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerows(data)

def fill_get_rename(right_format_file, right_order, index):
    # print(f'the index is {index}')
    data = read_csv(right_format_file)
    row = 10 + index
    start_column = 5
    reg_hours = float(get_reg_hours(data,row,start_column))
    # print(f' this is reg_hours {reg_hours}')
    # print(f'this is the right order before {right_order}')
    overtime = check_overtime(right_order[1],right_order[2],reg_hours)
    reg_hours = reg_hours/5/2

    right_order[3] = right_order[3] * reg_hours
    right_order[4] = right_order[4] * reg_hours
    right_order[5] = right_order[5] * reg_hours
    right_order[1] = overtime

    try:
        right_order[0] = float(right_order[0])
        overtime = float(overtime)
        overtime = round(overtime,2)
    except Exception as e:
        pass

    right_order[0] = right_order[0] - overtime

    

    del right_order[2]
    # print(f'this is the right order after {right_order}')

    for item in right_order:
        modify_cell(data, row,start_column,item)
        write_csv(right_format_file, data)
        data = read_csv(right_format_file)
        start_column = start_column + 1

    return right_order

def move_file(right_format_file):
    directory, filename = os.path.split(right_format_file)
    destination = os.path.join(directory, 'Filled')

    # Create the destination directory if it doesn't exist
    if not os.path.exists(destination):
        os.makedirs(destination)

    # Full destination path including the filename
    destination_file = os.path.join(destination, filename)

    # Check if the file already exists at the destination
    if os.path.exists(destination_file):
        os.remove(destination_file)  # Remove the existing file if it exists
        # print(f"Removed existing file: {destination_file}")
    
    # Move the file to the destination folder
    shutil.move(right_format_file, destination)
    # print(f"Moved file to: {destination}")

def move_file_check(right_format_file, folder_name):
    directory, filename = os.path.split(right_format_file)
    destination = os.path.join(directory, folder_name)

    # Create the destination directory if it doesn't exist
    if not os.path.exists(destination):
        os.makedirs(destination)

    # Full destination path including the filename
    destination_file = os.path.join(destination, filename)

    # Check if the file already exists at the destination
    if os.path.exists(destination_file):
        os.remove(destination_file)  # Remove the existing file if it exists
        print(f"Removed existing file: {destination_file}")

    # Move the file to the destination folder
    shutil.move(right_format_file, destination)
    # print(f"Moved file to: {destination}")



def get_reg_hours(data,row,start_column):
    reg_hours = data[row][start_column]
    return reg_hours

def check_overtime(week1, week2,reg_hours):
    max = reg_hours/2
    # print(f'week 1 {week1} and week2 {week2}')
    # print(f'max {max}')
    # print(f'week 1 type {type(week1)} and week2 type {type(week2)}')
    overtime1 = week1-max
    overtime2 = week2-max
    # print(f'overtime1 1 {overtime1} and overtime1 {overtime2}')
    if overtime1 > 0 and overtime2 > 0:
        overtime = overtime1+overtime2
        overtime = float(overtime)
        overtime = f"{overtime:.2f}"
        return overtime
    if overtime1 > 0:
        overtime1 = float(overtime1)
        overtime1 = f"{overtime1:.2f}"
        return overtime1
    if overtime2 > 0:
        overtime2 = float(overtime2)
        overtime2 = f"{overtime2:.2f}"
        return overtime2
    else:
        return 0

def rename_all(csv_path):
    if os.path.isdir(csv_path):
        for filename in os.listdir(csv_path):
            full_path = os.path.join(csv_path, filename)
            if filename.startswith('.') or not filename.lower().endswith('.csv'):
                continue
            place, date = get_rename(full_path)
            rename = place + ' ' + date + '.csv'
            directory, filename = os.path.split(full_path)
            old = full_path
            new = os.path.join(directory, rename)
            os.rename(old,new)

def get_rename(full_path):
    data = read_csv(full_path)
    renameplace = str(data[3][2])
    start_index = 10
    end_index = len(renameplace) - 12
    place = renameplace[start_index:end_index]
    renamedate = str(data[2][2])
    date = renamedate.replace('/2024', '').replace('/', '-')
    # print('date' + date)
    # print('place' + place)
    return place, date

def find_first_item(items):
    week1 = ['MON', 'TUE', 'TUES', 'WED', 'THU', 'FRI','SAT','SUN']
    for item in items:
        if item is not None:
            if item in week1:
                return item
    return None

def set_workdays(list):
    first_item = find_first_item(list)
    # print(f'irst_item {first_item}')
    newlist = []
    mon = 0
    for item in list:
        if item is None:
            pass
        if mon == 2:
            if isinstance(item,str):
                item = str(item) + '2'
                item = item.upper()
        if str(item).upper() == first_item:
            mon = mon + 1
            if mon == 2:
                if isinstance(item, str):
                    item = str(item) + '2'
                    item = item.upper()
        newlist.append(item)
    return newlist

def read_excel_ignore_hidden(file_path, sheet_name=None):
    # Load the workbook and select the sheet
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    # Collect visible columns and rows
    visible_rows = []
    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
        if not sheet.row_dimensions[row_idx].hidden:
            visible_rows.append([cell.value for cell in row])

    # Get visible columns from the first visible row
    if visible_rows:
        col_indices = [i for i, value in enumerate(visible_rows[0]) if value is not None]
        visible_data = [[row[i] for i in col_indices] for row in visible_rows]
        df = pd.DataFrame(visible_data[1:], columns=visible_data[0])
    else:
        df = pd.DataFrame()

    return df
